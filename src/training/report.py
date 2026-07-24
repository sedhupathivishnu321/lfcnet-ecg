"""
Generates every requested metric as both a PNG chart and an Excel workbook row/sheet:

| Metric                              | PNG                              | Excel sheet                         |
|--------------------------------------|-----------------------------------|--------------------------------------|
| Confusion matrix                     | plots/confusion_matrix.png       | Test Metrics (full matrix table)     |
| MCC                                   | plots/metrics_bar.png            | Test Metrics, Summary                |
| F1 curve (vs. threshold)             | plots/f1_curve.png               | - (image-only; peak value in Test Metrics) |
| Accuracy/Precision/Recall/F1         | plots/metrics_bar.png            | Test Metrics, Summary                |
| MSE                                   | -                                 | Test Metrics, Summary                |
| Training/validation loss curves      | plots/training_curves.png        | Training History (+ native line chart) |
| Efficiency (latency/FLOPs/params/size)| plots/efficiency.png            | Efficiency, Summary                  |

The Summary sheet's cells are live Excel formulas (=MAX, =INDEX/MATCH, cross-sheet
refs) - never hardcoded numbers - so it stays correct if you re-run training and
regenerate the workbook. If you edit this file and add new formulas, re-run
`python /mnt/skills/public/xlsx/scripts/recalc.py outputs/lafnet_report.xlsx`
(or open/save once in Excel/LibreOffice) so cached formula results populate.
"""
from __future__ import annotations

from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font

from src.config import ConfigNode
from src.training.metrics import f1_vs_threshold, roc_points

ARIAL = Font(name="Arial")
ARIAL_BOLD = Font(name="Arial", bold=True)


def _apply_arial(ws) -> None:
    for row in ws.iter_rows():
        for cell in row:
            cell.font = ARIAL_BOLD if cell.row == 1 else ARIAL


def _plot_confusion_matrix(cm: list[list[int]], out_path: Path) -> None:
    cm = np.array(cm)
    fig, ax = plt.subplots(figsize=(4, 4))
    ax.imshow(cm, cmap="Blues")
    for i in range(cm.shape[0]):
        for j in range(cm.shape[1]):
            ax.text(j, i, str(cm[i, j]), ha="center", va="center")
    ax.set_xticks([0, 1], labels=["Non-AF", "AF"])
    ax.set_yticks([0, 1], labels=["Non-AF", "AF"])
    ax.set_xlabel("Predicted")
    ax.set_ylabel("True")
    ax.set_title("Confusion Matrix")
    fig.tight_layout()
    fig.savefig(out_path, dpi=150)
    plt.close(fig)


def _plot_f1_curve(y_true, y_prob, out_path: Path) -> None:
    thresholds, f1s, best_t, best_f1 = f1_vs_threshold(np.array(y_true), np.array(y_prob))
    fig, ax = plt.subplots(figsize=(5, 4))
    ax.plot(thresholds, f1s)
    ax.axvline(best_t, color="red", linestyle="--", label=f"best t={best_t:.2f}, F1={best_f1:.2f}")
    ax.set_xlabel("Decision threshold")
    ax.set_ylabel("F1 score")
    ax.set_title("F1 vs. Threshold")
    ax.legend()
    fig.tight_layout()
    fig.savefig(out_path, dpi=150)
    plt.close(fig)


def _plot_metrics_bar(metrics: dict, out_path: Path) -> None:
    keys = ["accuracy", "precision", "recall", "f1", "mcc"]
    vals = [metrics[k] for k in keys]
    fig, ax = plt.subplots(figsize=(5, 4))
    ax.bar(keys, vals, color="#3b6fa0")
    ax.set_ylim(0, 1 if max(vals) <= 1 else max(vals) * 1.1)
    ax.set_title("Test Metrics")
    fig.tight_layout()
    fig.savefig(out_path, dpi=150)
    plt.close(fig)


def _plot_training_curves(history: list[dict], out_path: Path) -> None:
    if not history:
        return
    epochs = [h["epoch"] for h in history]
    fig, ax = plt.subplots(figsize=(6, 4))
    ax.plot(epochs, [h["train_loss"] for h in history], label="train_loss")
    ax.plot(epochs, [h["val_loss"] for h in history], label="val_loss")
    ax.set_xlabel("Epoch")
    ax.set_ylabel("Loss")
    ax.set_title("Training / Validation Loss")
    ax.legend()
    fig.tight_layout()
    fig.savefig(out_path, dpi=150)
    plt.close(fig)


def _plot_efficiency(efficiency: dict, out_path: Path) -> None:
    keys = ["params", "macs", "flops", "latency_ms_cpu"]
    labels = ["Params", "MACs", "FLOPs", "CPU Latency (ms)"]
    if efficiency.get("latency_ms_gpu") is not None:
        keys.append("latency_ms_gpu")
        labels.append("GPU Latency (ms)")
    vals = [efficiency[k] for k in keys]
    fig, ax = plt.subplots(figsize=(6, 4))
    ax.bar(labels, vals, color="#a05a3b")
    ax.set_title("Efficiency")
    ax.tick_params(axis="x", rotation=20)
    fig.tight_layout()
    fig.savefig(out_path, dpi=150)
    plt.close(fig)


def generate_report(cfg: ConfigNode, history: list[dict], report: dict) -> None:
    plots_dir = Path(cfg.paths.plots_dir)
    plots_dir.mkdir(parents=True, exist_ok=True)

    metrics = report["test_metrics"]
    efficiency = report["efficiency"]

    _plot_confusion_matrix(metrics["confusion_matrix"], plots_dir / "confusion_matrix.png")
    _plot_f1_curve(report["y_true"], report["y_prob"], plots_dir / "f1_curve.png")
    _plot_metrics_bar(metrics, plots_dir / "metrics_bar.png")
    _plot_training_curves(history, plots_dir / "training_curves.png")
    _plot_efficiency(efficiency, plots_dir / "efficiency.png")

    wb = Workbook()

    # --- Training History sheet ---
    ws_hist = wb.active
    ws_hist.title = "Training History"
    ws_hist.append(["epoch", "train_loss", "val_loss", "val_accuracy", "val_f1", "val_mcc", "val_roc_auc", "lr"])
    for h in history:
        ws_hist.append([h["epoch"], h["train_loss"], h["val_loss"], h["val_accuracy"], h["val_f1"], h["val_mcc"], h["val_roc_auc"], h["lr"]])
    _apply_arial(ws_hist)

    if history:
        chart = LineChart()
        chart.title = "Training / Validation Loss"
        chart.x_axis.title = "Epoch"
        chart.y_axis.title = "Loss"
        data = Reference(ws_hist, min_col=2, max_col=3, min_row=1, max_row=len(history) + 1)
        cats = Reference(ws_hist, min_col=1, min_row=2, max_row=len(history) + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws_hist.add_chart(chart, "J2")

    # --- Test Metrics sheet ---
    ws_test = wb.create_sheet("Test Metrics")
    ws_test.append(["metric", "value"])
    for k in ["accuracy", "precision", "recall", "f1", "mcc", "mse", "roc_auc"]:
        ws_test.append([k, metrics[k]])
    _, _, best_t, best_f1 = f1_vs_threshold(np.array(report["y_true"]), np.array(report["y_prob"]))
    ws_test.append(["f1_curve_best_threshold", best_t])
    ws_test.append(["f1_curve_best_f1", best_f1])
    ws_test.append([])
    ws_test.append(["confusion_matrix", "pred_non_af", "pred_af"])
    cm = metrics["confusion_matrix"]
    ws_test.append(["true_non_af", cm[0][0], cm[0][1]])
    ws_test.append(["true_af", cm[1][0], cm[1][1]])
    _apply_arial(ws_test)

    # --- Efficiency sheet ---
    ws_eff = wb.create_sheet("Efficiency")
    ws_eff.append(["metric", "value"])
    for k, v in efficiency.items():
        ws_eff.append([k, v])
    _apply_arial(ws_eff)

    # --- Confidence Intervals sheet (bootstrap, from src.training.metrics.bootstrap_ci_all_metrics) ---
    ci = report.get("confidence_intervals")
    if ci:
        ws_ci = wb.create_sheet("Confidence Intervals")
        ws_ci.append(["metric", "point_estimate", "ci_lower", "ci_upper", "confidence_level", "n_bootstrap"])
        for metric_name, vals in ci.items():
            ws_ci.append([
                metric_name, vals["point_estimate"], vals["ci_lower"], vals["ci_upper"],
                vals["confidence_level"], vals["n_iterations"],
            ])
        _apply_arial(ws_ci)

    # --- Summary sheet: live formulas only, never hardcoded numbers ---
    ws_sum = wb.create_sheet("Summary")
    ws_sum.append(["Model", cfg.model.name])
    ws_sum.append(["Best Val F1 (over all epochs)", "=MAX('Training History'!E:E)"])
    ws_sum.append(["Best Val MCC (over all epochs)", "=MAX('Training History'!F:F)"])
    ws_sum.append(["Epoch of Best Val F1",
                    "=INDEX('Training History'!A:A,MATCH(MAX('Training History'!E:E),'Training History'!E:E,0))"])
    ws_sum.append(["Test Accuracy", "=INDEX('Test Metrics'!B:B,MATCH(\"accuracy\",'Test Metrics'!A:A,0))"])
    ws_sum.append(["Test Precision", "=INDEX('Test Metrics'!B:B,MATCH(\"precision\",'Test Metrics'!A:A,0))"])
    ws_sum.append(["Test Recall", "=INDEX('Test Metrics'!B:B,MATCH(\"recall\",'Test Metrics'!A:A,0))"])
    ws_sum.append(["Test F1", "=INDEX('Test Metrics'!B:B,MATCH(\"f1\",'Test Metrics'!A:A,0))"])
    ws_sum.append(["Test MCC", "=INDEX('Test Metrics'!B:B,MATCH(\"mcc\",'Test Metrics'!A:A,0))"])
    ws_sum.append(["Test MSE", "=INDEX('Test Metrics'!B:B,MATCH(\"mse\",'Test Metrics'!A:A,0))"])
    ws_sum.append(["Test ROC-AUC", "=INDEX('Test Metrics'!B:B,MATCH(\"roc_auc\",'Test Metrics'!A:A,0))"])
    ws_sum.append(["Params", "=INDEX(Efficiency!B:B,MATCH(\"params\",Efficiency!A:A,0))"])
    ws_sum.append(["MACs", "=INDEX(Efficiency!B:B,MATCH(\"macs\",Efficiency!A:A,0))"])
    ws_sum.append(["FLOPs", "=INDEX(Efficiency!B:B,MATCH(\"flops\",Efficiency!A:A,0))"])
    ws_sum.append(["CPU Latency (ms)", "=INDEX(Efficiency!B:B,MATCH(\"latency_ms_cpu\",Efficiency!A:A,0))"])
    ws_sum.append(["GPU Latency (ms)", "=INDEX(Efficiency!B:B,MATCH(\"latency_ms_gpu\",Efficiency!A:A,0))"])
    ws_sum.append(["Energy per inference (uJ)", "=INDEX(Efficiency!B:B,MATCH(\"energy_uj_per_inference\",Efficiency!A:A,0))"])
    if ci:
        ws_sum.append(["F1 95% CI lower",
                        "=INDEX('Confidence Intervals'!C:C,MATCH(\"f1\",'Confidence Intervals'!A:A,0))"])
        ws_sum.append(["F1 95% CI upper",
                        "=INDEX('Confidence Intervals'!D:D,MATCH(\"f1\",'Confidence Intervals'!A:A,0))"])
    _apply_arial(ws_sum)

    for ws in wb.worksheets:
        ws.column_dimensions["A"].width = 32
        ws.column_dimensions["B"].width = 20

    out_path = Path(cfg.paths.report_xlsx)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"[report] Workbook -> {out_path}")
    print("[report] Formulas are uncached until recalculated - run "
          "`python /mnt/skills/public/xlsx/scripts/recalc.py " + str(out_path) + "` "
          "or open/save once in Excel/LibreOffice.")
